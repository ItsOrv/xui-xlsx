#!/usr/bin/env python3
"""
Export x-ui panel users from backup .db files to CSV and Excel.
Columns: source file, inbound, email (remark), UUID, used/remaining traffic (GB), total limit (GB).
"""

import argparse
import csv
import json
import sqlite3
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

BYTES_PER_GB = 1024**3
DEFAULT_OUTPUT_EXCEL = Path(__file__).parent / "xui_users_export.xlsx"
DEFAULT_OUTPUT_CSV = Path(__file__).parent / "xui_users_export.csv"
HEADERS = [
    "Source",
    "Inbound",
    "Email",
    "UUID",
    "Used (GB)",
    "Remaining (GB)",
    "Total limit (GB)",
]


@dataclass(slots=True)
class UserTraffic:
    source: str
    inbound: str
    email: str
    uuid: str
    used_gb: float
    remaining_gb: Optional[float]
    total_gb: Optional[float]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export x-ui users from one or more SQLite backups."
    )
    parser.add_argument(
        "--db",
        action="append",
        dest="db_files",
        default=[],
        help="Path to a backup DB file (can be used multiple times).",
    )
    parser.add_argument(
        "--input-dir",
        default=str(Path(__file__).parent),
        help="Directory to scan for DB files if --db is not provided.",
    )
    parser.add_argument(
        "--pattern",
        default="*.db",
        help="Glob pattern used with --input-dir (default: *.db).",
    )
    parser.add_argument(
        "--csv-output",
        default=str(DEFAULT_OUTPUT_CSV),
        help="CSV output path.",
    )
    parser.add_argument(
        "--xlsx-output",
        default=str(DEFAULT_OUTPUT_EXCEL),
        help="Excel output path.",
    )
    parser.add_argument(
        "--no-csv",
        action="store_true",
        help="Disable CSV output.",
    )
    parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="Disable XLSX output.",
    )
    parser.add_argument(
        "--email-contains",
        default="",
        help="Keep users where email contains this text (case-insensitive).",
    )
    parser.add_argument(
        "--min-remaining",
        type=float,
        default=None,
        help="Keep users with remaining GB >= this value (unlimited excluded).",
    )
    parser.add_argument(
        "--max-remaining",
        type=float,
        default=None,
        help="Keep users with remaining GB <= this value (unlimited excluded).",
    )
    parser.add_argument(
        "--only-unlimited",
        action="store_true",
        help="Keep only unlimited users.",
    )
    parser.add_argument(
        "--sort-by",
        choices=["remaining", "used", "email", "source", "total"],
        default="remaining",
        help="Sort key inside each inbound group.",
    )
    parser.add_argument(
        "--asc",
        action="store_true",
        help="Sort ascending inside each inbound group.",
    )
    return parser.parse_args()


def resolve_db_paths(args: argparse.Namespace) -> list[Path]:
    if args.db_files:
        return [
            Path(p).expanduser().resolve()
            for p in args.db_files
            if Path(p).expanduser().suffix.lower() == ".db"
        ]
    input_dir = Path(args.input_dir).expanduser().resolve()
    if not input_dir.exists() or not input_dir.is_dir():
        print(f"Input directory not found: {input_dir}")
        return []
    return sorted(input_dir.glob(args.pattern))


def get_inbound_clients_by_email(conn: sqlite3.Connection) -> dict[int, dict[str, str]]:
    """inbound_id -> { email: uuid } from inbounds.settings JSON."""
    cur = conn.execute(
        "SELECT id, settings FROM inbounds WHERE settings IS NOT NULL AND settings != ''"
    )
    out: dict[int, dict[str, str]] = {}
    for inbound_id, settings_str in cur.fetchall():
        try:
            data = json.loads(settings_str)
            clients = data.get("clients") or []
            out[inbound_id] = {c.get("email", ""): c.get("id", "") for c in clients if c}
        except (json.JSONDecodeError, TypeError, AttributeError):
            out[inbound_id] = {}
    return out


def get_inbound_labels(conn: sqlite3.Connection) -> dict[int, str]:
    """inbound_id -> readable inbound label."""
    cur = conn.execute("SELECT id, remark, tag, protocol FROM inbounds")
    labels: dict[int, str] = {}
    for inbound_id, remark, tag, protocol in cur.fetchall():
        parts = [
            str(part).strip()
            for part in (remark, tag, protocol)
            if part and str(part).strip() and str(part).strip() != "-"
        ]
        info = " | ".join(parts)
        labels[inbound_id] = f"{inbound_id} - {info}" if info else str(inbound_id)
    return labels


def bytes_to_gb(value: int) -> float:
    return round(value / BYTES_PER_GB, 2)


def load_db(path: Path) -> tuple[list[UserTraffic], str | None]:
    """
    Return (rows, error_message).
    rows: list of UserTraffic records.
    """
    if not path.exists():
        return [], f"File not found: {path}"

    try:
        conn = sqlite3.connect(str(path))
    except sqlite3.Error as exc:
        return [], f"Could not open DB: {path.name} ({exc})"

    conn.row_factory = sqlite3.Row
    try:
        try:
            email_to_uuid = get_inbound_clients_by_email(conn)
            inbound_labels = get_inbound_labels(conn)
            rows = conn.execute(
                """
                SELECT inbound_id, email, up, down, total
                FROM client_traffics
                WHERE email IS NOT NULL AND email != ''
                """
            ).fetchall()
        except sqlite3.OperationalError as exc:
            return [], f"{path.name}: missing expected schema/tables ({exc})"

        result: list[UserTraffic] = []
        for r in rows:
            inbound_id = r["inbound_id"]
            email = (r["email"] or "").strip()
            up = int(r["up"] or 0)
            down = int(r["down"] or 0)
            total_bytes = int(r["total"] or 0)
            uuid = (email_to_uuid.get(inbound_id) or {}).get(email, "")
            inbound_label = inbound_labels.get(inbound_id, str(inbound_id))

            if total_bytes > 0:
                used_bytes = max(0, up + down)
                remaining_bytes = max(0, total_bytes - up - down)
                used_gb = bytes_to_gb(used_bytes)
                remaining_gb = bytes_to_gb(remaining_bytes)
                total_gb = bytes_to_gb(total_bytes)
            else:
                used_gb = bytes_to_gb(up + down)
                remaining_gb = None
                total_gb = None

            result.append(
                UserTraffic(
                    source=path.name,
                    inbound=inbound_label,
                    email=email,
                    uuid=uuid,
                    used_gb=used_gb,
                    remaining_gb=remaining_gb,
                    total_gb=total_gb,
                )
            )
        return result, None
    finally:
        conn.close()


def apply_filters(rows: list[UserTraffic], args: argparse.Namespace) -> list[UserTraffic]:
    out = rows

    if args.email_contains:
        q = args.email_contains.lower().strip()
        out = [r for r in out if q in r.email.lower()]

    if args.only_unlimited:
        out = [r for r in out if r.remaining_gb is None]
    else:
        if args.min_remaining is not None:
            out = [
                r
                for r in out
                if r.remaining_gb is not None and r.remaining_gb >= args.min_remaining
            ]
        if args.max_remaining is not None:
            out = [
                r
                for r in out
                if r.remaining_gb is not None and r.remaining_gb <= args.max_remaining
            ]
    return out


def numeric_key(value: Optional[float], asc: bool) -> tuple[int, float]:
    if value is None:
        return (1, 0.0)
    return (0, value if asc else -value)


def sort_group(group: list[UserTraffic], args: argparse.Namespace) -> list[UserTraffic]:
    if args.sort_by == "email":
        return sorted(group, key=lambda r: r.email.lower(), reverse=not args.asc)
    if args.sort_by == "source":
        return sorted(
            group,
            key=lambda r: (r.source.lower(), r.email.lower()),
            reverse=not args.asc,
        )
    if args.sort_by == "used":
        return sorted(
            group,
            key=lambda r: (numeric_key(r.used_gb, args.asc), r.email.lower()),
        )
    if args.sort_by == "total":
        return sorted(
            group,
            key=lambda r: (numeric_key(r.total_gb, args.asc), r.email.lower()),
        )
    return sorted(
        group,
        key=lambda r: (numeric_key(r.remaining_gb, args.asc), r.email.lower()),
    )


def sort_rows(rows: list[UserTraffic], args: argparse.Namespace) -> list[UserTraffic]:
    grouped: dict[str, list[UserTraffic]] = {}
    for row in rows:
        grouped.setdefault(row.inbound, []).append(row)

    sorted_rows: list[UserTraffic] = []
    for inbound in sorted(grouped.keys(), key=str.lower):
        sorted_rows.extend(sort_group(grouped[inbound], args))
    return sorted_rows


def row_values(r: UserTraffic) -> list:
    return [
        r.source,
        r.inbound,
        r.email,
        r.uuid,
        r.used_gb,
        r.remaining_gb if r.remaining_gb is not None else "Unlimited",
        r.total_gb if r.total_gb is not None else "Unlimited",
    ]


def write_csv(rows: Iterable[UserTraffic], output_csv: Path) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(row_values(r) for r in rows)


def write_xlsx(rows: Iterable[UserTraffic], output_excel: Path) -> None:
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(rows, 2):
        for col_idx, val in enumerate(row_values(r), 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24
    ws.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 36
    wb.save(output_excel)


def print_summary(
    rows: list[UserTraffic], per_db_counts: dict[str, int], db_errors: list[str]
) -> None:
    unlimited_count = sum(1 for r in rows if r.remaining_gb is None)
    missing_uuid_count = sum(1 for r in rows if not r.uuid)

    print("\nSummary:")
    print(f"- Total exported users: {len(rows)}")
    print(f"- Unlimited users: {unlimited_count}")
    print(f"- Users without UUID: {missing_uuid_count}")
    print("- Per-source counts:")
    for source, count in sorted(per_db_counts.items()):
        print(f"  - {source}: {count}")

    if db_errors:
        print("- Warnings:")
        for err in db_errors:
            print(f"  - {err}")


def main() -> int:
    args = parse_args()

    if args.no_csv and args.no_xlsx:
        print("Nothing to do: both CSV and XLSX outputs are disabled.")
        return 2

    db_paths = resolve_db_paths(args)
    if not db_paths:
        print("No database files found.")
        return 1

    all_rows: list[UserTraffic] = []
    per_db_counts: dict[str, int] = {}
    db_errors: list[str] = []

    for db_path in db_paths:
        rows, err = load_db(db_path)
        if err:
            db_errors.append(err)
            print(f"Skipped {db_path.name}: {err}")
            continue
        all_rows.extend(rows)
        per_db_counts[db_path.name] = len(rows)
        print(f"Loaded {len(rows)} users from {db_path.name}")

    if not all_rows:
        print("No users found in readable DB files.")
        if db_errors:
            print("Use warnings above to fix DB/schema issues.")
        return 1

    filtered_rows = apply_filters(all_rows, args)
    if not filtered_rows:
        print("No users matched the selected filters.")
        return 1

    sorted_rows = sort_rows(filtered_rows, args)

    output_csv = Path(args.csv_output).expanduser().resolve()
    output_excel = Path(args.xlsx_output).expanduser().resolve()

    if not args.no_csv:
        write_csv(sorted_rows, output_csv)
        print(f"Saved {len(sorted_rows)} users to {output_csv}")

    if not args.no_xlsx:
        if HAS_OPENPYXL:
            write_xlsx(sorted_rows, output_excel)
            print(f"Saved {len(sorted_rows)} users to {output_excel}")
        else:
            print("XLSX skipped: openpyxl is not installed (pip install openpyxl).")

    print_summary(sorted_rows, per_db_counts, db_errors)
    return 0


if __name__ == "__main__":
    sys.exit(main())
